from openpyxl import load_workbook

import datetime
import logging


def get_template(filename='template.xlsx'):
    excel_logger = logging.getLogger("generator.excel")

    try:
        wb = load_workbook(filename)
        sheet = wb['template']
    except KeyError:
        excel_logger.exception('Tab must have name: "template".')
    except FileNotFoundError:
        excel_logger.exception('File ' + filename + ' not found.')
    policies_template = {}
    task_name = "default_task"
    for i in range(3, sheet.max_row + 1):
        if sheet['A' + str(i)].value:
            task_name = sheet['A' + str(i)].value

        policies_template.update({i: {'task': task_name, 'filename': sheet['D' + str(i)].value,
                                      'sender': sheet['E' + str(i)].value,
                                      'recipient': sheet['F' + str(i)].value,
                                      'policy_name': sheet['G' + str(i)].value if sheet['G' + str(i)].value else ' ',
                                      'verdict': sheet['I' + str(i)].value,
                                      'violation_level': sheet['J' + str(i)].value,
                                      'tag': sheet['K' + str(i)].value if sheet['K' + str(i)].value else ' '}})

    wb.close()

    return policies_template
